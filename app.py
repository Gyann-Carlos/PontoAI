import streamlit as st
import csv
import json
import smtplib
import os
import socket
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from urllib.request import Request, urlopen
from urllib.error import URLError
from io import StringIO, BytesIO
from collections import defaultdict
import pandas as pd

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL = "llama3.1:8b"
CSV_PADRAO = "folha_ponto.csv"

TIPOS_IA = ["Local (Ollama)", "Online (API)"]


def listar_modelos_ollama():
    try:
        req = Request("http://localhost:11434/api/tags")
        with urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
            return [m["name"] for m in data.get("models", [])]
    except Exception:
        return []


def consultar_ollama(prompt, modelo):
    payload = json.dumps({"model": modelo, "prompt": prompt, "stream": False}).encode("utf-8")
    req = Request(OLLAMA_URL, data=payload, headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=300) as resp:
            return json.loads(resp.read())["response"]
    except URLError as e:
        return f"Erro ao conectar no Ollama: {e.reason}"
    except Exception as e:
        return f"Erro: {e}"


def consultar_api_online(prompt, api_key, endpoint, modelo):
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    payload = json.dumps({
        "model": modelo,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
    }).encode("utf-8")
    try:
        req = Request(endpoint, data=payload, headers=headers)
        with urlopen(req, timeout=300) as resp:
            data = json.loads(resp.read())
            return data.get("choices", [{}])[0].get("message", {}).get("content", "")
    except URLError as e:
        return f"Erro ao conectar na API: {e.reason}"
    except Exception as e:
        return f"Erro: {e}"


def consultar_ia(prompt, tipo_ia, modelo, api_key="", endpoint=""):
    if tipo_ia == "Local (Ollama)":
        return consultar_ollama(prompt, modelo)
    else:
        return consultar_api_online(prompt, api_key, endpoint, modelo)


def fmt_hora(valor):
    if not valor or not valor.strip():
        return "--"
    h, m = valor.strip().split(":")
    return f"{h}:{m}:00"

def fmt_saldo(dec):
    if dec == 0:
        return "00:00:00"
    sinal = "+" if dec > 0 else "-"
    total_seg = int(abs(dec) * 3600)
    h = total_seg // 3600
    m = (total_seg % 3600) // 60
    s = total_seg % 60
    return f"{sinal}{h:02d}:{m:02d}:{s:02d}"

def parse_hora(valor):
    if not valor or not str(valor).strip():
        return None
    return datetime.strptime(str(valor).strip(), "%H:%M")


def normalizar_colunas(rows):
    mapa = {
        "matricula": "colaborador_id",
        "matrícula": "colaborador_id",
        "colaborador_id": "colaborador_id",
        "id": "colaborador_id",
        "nome": "nome",
        "nome completo": "nome",
        "data": "data",
        "dia": "data",
        "date": "data",
        "entrada": "entrada",
        "hor entrada": "entrada",
        "hora entrada": "entrada",
        "horário entrada": "entrada",
        "saida_almoco": "saida_almoco",
        "saida almoço": "saida_almoco",
        "saída almoço": "saida_almoco",
        "retorno_almoco": "retorno_almoco",
        "retorno almoço": "retorno_almoco",
        "volta almoço": "retorno_almoco",
        "saida": "saida",
        "saída": "saida",
        "hor saída": "saida",
        "hor saida": "saida",
        "hora saída": "saida",
        "marc saida": "saida",
        "marc saída": "saida",
        "cargo": "cargo",
        "carga_horaria_diaria": "carga_horaria_diaria",
        "carga horária": "carga_horaria_diaria",
        "carga horaria": "carga_horaria_diaria",
        "qtde de horas": "carga_horaria_diaria",
        "email": "email",
        "e-mail": "email",
        "ocorrencia": "ocorrencia",
        "ocorrência": "ocorrencia",
    }
    normalizados = []
    for row in rows:
        novo = {}
        for chave, valor in row.items():
            chave_limpa = chave.strip().lower() if chave else ""
            destino = mapa.get(chave_limpa, chave_limpa)
            if destino not in novo or not novo.get(destino):
                novo[destino] = valor
        if "data" in novo:
            val = novo["data"]
            if isinstance(val, datetime):
                novo["data"] = val.strftime("%d/%m/%Y")
            elif isinstance(val, str) and val.strip():
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
                    try:
                        novo["data"] = datetime.strptime(val.strip(), fmt).strftime("%d/%m/%Y")
                        break
                    except ValueError:
                        continue
        for campo in ("entrada", "saida", "saida_almoco", "retorno_almoco"):
            if campo in novo:
                if isinstance(novo[campo], datetime):
                    novo[campo] = novo[campo].strftime("%H:%M")
                else:
                    val = str(novo[campo]).strip().lstrip("'").strip()
                    novo[campo] = val if val else ""
        if "carga_horaria_diaria" in novo:
            val = str(novo["carga_horaria_diaria"]).replace(",", ".").strip().lstrip("'")
            try:
                if ":" in val:
                    h, m = val.split(":")
                    novo["carga_horaria_diaria"] = str(int(h) + int(int(m) / 60))
                else:
                    novo["carga_horaria_diaria"] = str(int(float(val)))
            except (ValueError, ZeroDivisionError):
                novo["carga_horaria_diaria"] = "8"
        else:
            novo["carga_horaria_diaria"] = "8"
        if "cargo" not in novo or not novo.get("cargo"):
            novo["cargo"] = "Colaborador"
        if "saida_almoco" not in novo:
            novo["saida_almoco"] = ""
        if "retorno_almoco" not in novo:
            novo["retorno_almoco"] = ""
        normalizados.append(novo)
    return normalizados


def ler_csv(arquivo):
    raw = arquivo.read()
    for enc in ("utf-8-sig", "latin-1", "cp1252", "iso-8859-1"):
        try:
            content = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        content = raw.decode("latin-1", errors="replace")
    return list(csv.DictReader(StringIO(content), delimiter=";"))


def ler_xlsx(arquivo):
    if not openpyxl:
        raise ImportError("openpyxl nao instalado. Instale com: pip install openpyxl")
    wb = openpyxl.load_workbook(arquivo)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(headers, row):
            chave = str(h or "")
            if isinstance(v, datetime):
                d[chave] = v
            else:
                d[chave] = str(v or "")
        rows.append(d)
    return rows


def ler_pdf(arquivo):
    if not pdfplumber:
        raise ImportError("pdfplumber nao instalado. Instale com: pip install pdfplumber")
    rows = []
    with pdfplumber.open(arquivo) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                headers = [str(h or "").strip() for h in table[0]]
                for data_row in table[1:]:
                    row = {}
                    for h, v in zip(headers, data_row):
                        row[h] = str(v or "").strip()
                    rows.append(row)
    return rows


def ler_arquivo(arquivo, nome_arquivo):
    ext = os.path.splitext(nome_arquivo)[1].lower()
    if ext == ".csv":
        rows = ler_csv(arquivo)
    elif ext in (".xlsx", ".xls"):
        rows = ler_xlsx(arquivo)
    elif ext == ".pdf":
        rows = ler_pdf(arquivo)
    else:
        raise ValueError(f"Formato nao suportado: {ext}")
    return normalizar_colunas(rows)


def calcular_horas_trabalhadas(entrada, saida_almoco, retorno_almoco, saida):
    if not entrada or not saida:
        return 0.0
    e = parse_hora(entrada)
    sm = parse_hora(saida_almoco)
    rm = parse_hora(retorno_almoco)
    s = parse_hora(saida)
    total = (s - e).total_seconds() / 3600
    if sm and rm:
        total -= (rm - sm).total_seconds() / 3600
    return round(total, 2)


def processar_registros(rows):
    colaboradores = {}
    tem_email = "email" in rows[0] if rows else False
    emails_por_colab = {}
    if tem_email:
        for row in rows:
            cid = row["colaborador_id"]
            email_raw = row.get("email", "").strip()
            if email_raw and cid not in emails_por_colab:
                emails_por_colab[cid] = email_raw
    campos_ponto = ["entrada", "saida"]
    tem_almoco = any(row.get("saida_almoco", "").strip() for row in rows)
    if tem_almoco:
        campos_ponto += ["saida_almoco", "retorno_almoco"]
    for row in rows:
        cid = row["colaborador_id"]
        if cid not in colaboradores:
            email = emails_por_colab.get(cid, "")
            if not email:
                nome_clean = row["nome"].lower().replace(" ", ".").replace("�", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ç", "c").replace("ã", "a").replace("õ", "o").replace("ê", "e").replace("ô", "o")
                email = f"{nome_clean}@positivosmais.com"
            colaboradores[cid] = {
                "nome": row["nome"],
                "email": email,
                "cargo": row.get("cargo", "Colaborador"),
                "carga_diaria": float(row.get("carga_horaria_diaria", "8")),
                "dias": [],
            }
        problemas = []
        ocorrencia = row.get("ocorrencia", "").strip()
        if ocorrencia:
            mapa_ocorrencia = {
                "s/marcação de entrada": "Entrada",
                "s/marcação de saída": "Saída",
                "entrada em atraso": "Entrada em atraso",
                "saída antecipada": "Saída antecipada",
                "intervalo irregular": "Intervalo irregular",
                "marcação irregular": "Marcação irregular",
            }
            for chave, rotulo in mapa_ocorrencia.items():
                if ocorrencia.lower().startswith(chave):
                    problemas.append(rotulo)
        for campo in campos_ponto:
            rotulo_map = {
                "entrada": "Entrada",
                "saida": "Saída",
                "saida_almoco": "Saida almoco",
                "retorno_almoco": "Retorno almoco",
            }
            if not row.get(campo, "").strip():
                rotulo = rotulo_map[campo]
                if rotulo not in problemas:
                    problemas.append(rotulo)

        horas_trab = calcular_horas_trabalhadas(
            row.get("entrada", ""),
            row.get("saida_almoco", "") if tem_almoco else "",
            row.get("retorno_almoco", "") if tem_almoco else "",
            row.get("saida", ""),
        )
        carga = colaboradores[cid]["carga_diaria"]
        saldo = round(horas_trab - carga, 2)

        colaboradores[cid]["dias"].append({
            "data": row.get("data", ""),
            "entrada": row.get("entrada", ""),
            "saida_almoco": row.get("saida_almoco", ""),
            "retorno_almoco": row.get("retorno_almoco", ""),
            "saida": row.get("saida", ""),
            "horas_trabalhadas": horas_trab,
            "saldo": saldo,
            "problemas": problemas,
        })
    return colaboradores


def gerar_resumo_texto(colaboradores):
    linhas = []
    for cid, info in colaboradores.items():
        linhas.append(f"Colaborador: {info['nome']} (ID {cid}) - {info['cargo']} - Carga diaria: {info['carga_diaria']}h")
        saldo_total = sum(d["saldo"] for d in info["dias"])
        incompletos = [d for d in info["dias"] if d["problemas"]]
        if incompletos:
            linhas.append("  Dias com registro incompleto:")
            for d in incompletos:
                linhas.append(f"  - {d['data']}: {', '.join(d['problemas'])}")
        else:
            linhas.append("  Todos os dias com registro completo.")
        if saldo_total > 0:
            linhas.append(f"  Saldo: {fmt_saldo(saldo_total)} (horas extras)")
        elif saldo_total < 0:
            linhas.append(f"  Saldo: {fmt_saldo(saldo_total)} (horas devendo)")
        else:
            linhas.append("  Saldo: 0h (dentro do esperado)")
        linhas.append("")
    return "\n".join(linhas)


def gerar_email_ia(nome, pendencia, dias_problema, saldo_total, cargo, tipo_ia, modelo, api_key="", endpoint=""):
    prompt = f"""Gere um e-mail profissional e cordial em portugues para um colaborador que esta com pendencias na folha de ponto.

Colaborador: {nome}
Cargo: {cargo}
Tipo de pendencia: {pendencia}
Dias com problemas: {dias_problema}
Saldo total de horas: {saldo_total}

O e-mail deve:
- Ser educado e profissional
- Explicar claramente qual a pendencia
- Solicitar que o colaborador regularize a situacao
- Ter um assunto claro
- Ser assinado como "RH / Departamento Pessoal"

Responda APENAS com o e-mail completo (assunto + corpo)."""
    return consultar_ia(prompt, tipo_ia, modelo, api_key, endpoint)


def enviar_email_smtp(assunto, corpo, destino, smtp_host, smtp_port, smtp_user, smtp_pass, smtp_dest):
    msg = MIMEMultipart()
    msg["From"] = smtp_dest
    msg["To"] = destino
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo, "plain", "utf-8"))
    try:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        return True, None
    except socket.gaierror:
        return False, f"Nao foi possivel resolver o servidor SMTP '{smtp_host}'. Verifique se o host esta correto."
    except smtplib.SMTPAuthenticationError:
        return False, "Falha de autenticacao. Verifique usuario e senha."
    except smtplib.SMTPConnectError:
        return False, f"Nao foi possivel conectar em {smtp_host}:{smtp_port}. Verifique host e porta."
    except smtplib.SMTPException as e:
        return False, f"Erro SMTP: {e}"
    except Exception as e:
        return False, str(e)


st.set_page_config(page_title="Análise de Folha de Ponto", layout="wide")
st.title("📊 Análise de Folha de Ponto")
st.markdown("---")

with st.sidebar:
    st.header("Configuração")
    arquivo = st.file_uploader("Upload do arquivo", type=["csv", "xlsx", "xls", "pdf"])
    if arquivo:
        try:
            rows = ler_arquivo(arquivo, arquivo.name)
            st.success(f"Arquivo carregado: {arquivo.name} ({len(rows)} registros)")
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            rows = []
    else:
        try:
            for nome in [CSV_PADRAO, "folha_ponto2.csv", "AJUSTES 19 DE JUNHO.xlsx"]:
                if os.path.exists(nome):
                    with open(nome, "rb") as f:
                        buf = BytesIO(f.read())
                    rows = ler_arquivo(buf, nome)
                    st.info(f"Usando arquivo padrão: {nome}")
                    break
            else:
                st.error("Nenhum arquivo encontrado. Faça upload de um CSV, XLSX ou PDF.")
                st.stop()
        except Exception as e:
            st.error(f"Erro ao ler arquivo padrão: {e}")
            st.stop()

    if not rows:
        st.warning("Nenhum dado encontrado no arquivo.")
        st.stop()

    if st.button("🔄 Processar dados"):
        st.session_state.processado = True

    st.markdown("---")
    st.header("IA")
    tipo_ia = st.radio("Tipo de IA", TIPOS_IA, key="tipo_ia")
    modelo_ia = MODEL
    api_key_ia = ""
    endpoint_ia = ""
    if tipo_ia == "Local (Ollama)":
        modelos = listar_modelos_ollama()
        if not modelos:
            modelos = [MODEL]
        modelo_ia = st.selectbox("Modelo Ollama", modelos, index=modelos.index(MODEL) if MODEL in modelos else 0)
    else:
        endpoint_ia = st.text_input("Endpoint API", value="https://api.openai.com/v1/chat/completions")
        api_key_ia = st.text_input("API Key", type="password")
        modelo_ia = st.text_input("Modelo", value="gpt-4o-mini")
    if st.button("🤖 Analisar com IA"):
        st.session_state.analisar_ia = True

    st.markdown("---")
    st.header("Filtro de Data")
    datas = sorted(set(r["data"] for r in rows))
    if datas:
        data_min = datetime.strptime(datas[0], "%d/%m/%Y")
        data_max = datetime.strptime(datas[-1], "%d/%m/%Y")
        filtro_inicio = st.date_input("Data inicial", value=data_min, min_value=data_min, max_value=data_max, format="DD/MM/YYYY")
        filtro_fim = st.date_input("Data final", value=data_max, min_value=data_min, max_value=data_max, format="DD/MM/YYYY")
    else:
        filtro_inicio = None
        filtro_fim = None

    st.markdown("---")
    st.header("Envio de E-mail")
    usar_smtp = st.checkbox("Configurar SMTP")
    if usar_smtp:
        smtp_host = st.text_input("SMTP Host", value="smtp.positivosmais.com")
        smtp_port = st.number_input("SMTP Port", value=587)
        smtp_user = st.text_input("Usuário (e-mail)")
        smtp_pass = st.text_input("Senha", type="password")
        smtp_dest = st.text_input("E-mail do remetente (From)")
    else:
        smtp_host = smtp_port = smtp_user = smtp_pass = smtp_dest = None
    if st.button("✉️ Enviar alertas por e-mail"):
        st.session_state.enviar_email = True

if "processado" not in st.session_state:
    st.session_state.processado = False
if "analisar_ia" not in st.session_state:
    st.session_state.analisar_ia = False
if "enviar_email" not in st.session_state:
    st.session_state.enviar_email = False

colaboradores = processar_registros(rows)

def filtrar_por_data(colabs, inicio, fim):
    if not inicio or not fim:
        return colabs
    filtrados = {}
    for cid, info in colabs.items():
        dias_filtrados = []
        for d in info["dias"]:
            dt = datetime.strptime(d["data"], "%d/%m/%Y")
            if inicio <= dt.date() <= fim:
                dias_filtrados.append(d)
        if dias_filtrados:
            filtrados[cid] = {**info, "dias": dias_filtrados}
    return filtrados

if filtro_inicio and filtro_fim:
    colaboradores = filtrar_por_data(colaboradores, filtro_inicio, filtro_fim)

if st.session_state.processado or True:
    df_all = []
    df_problemas = []
    df_saldo = []

    for cid, info in colaboradores.items():
        saldo_total = sum(d["saldo"] for d in info["dias"])
        df_saldo.append({
            "ID": cid,
            "Nome": info["nome"],
            "Cargo": info["cargo"],
            "Saldo Total": fmt_saldo(saldo_total),
            "Situação": "Horas extras" if saldo_total > 0 else ("Devendo horas" if saldo_total < 0 else "OK"),
        })
        for d in info["dias"]:
            df_all.append({
                "ID": cid,
                "Nome": info["nome"],
                "Data": d["data"],
                "Entrada": fmt_hora(d["entrada"]),
                "Saída Almoço": fmt_hora(d["saida_almoco"]),
                "Retorno Almoço": fmt_hora(d["retorno_almoco"]),
                "Saída": fmt_hora(d["saida"]),
                "Horas Trab.": d["horas_trabalhadas"],
                "Saldo": fmt_saldo(d["saldo"]),
            })
            if d["problemas"]:
                df_problemas.append({
                    "ID": cid,
                    "Nome": info["nome"],
                    "Data": d["data"],
                    "Campos faltando": ", ".join(d["problemas"]),
                })

    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Visão Geral",
        "⚠️ Registros Incompletos",
        "💰 Saldo de Horas",
        "🔍 Detalhamento",
    ])

    with tab1:
        col1, col2, col3, col4 = st.columns(4)
        total_colab = len(colaboradores)
        total_dias = len(df_all)
        total_problemas = len(df_problemas)
        devendo = sum(1 for s in df_saldo if s["Situação"] == "Devendo horas")
        with col1:
            st.metric("Colaboradores", total_colab)
        with col2:
            st.metric("Total de registros", total_dias)
        with col3:
            st.metric("Registros incompletos", total_problemas)
        with col4:
            st.metric("Colabs devendo horas", devendo, delta_color="inverse")

        st.subheader("Resumo por colaborador")
        df_saldo_df = pd.DataFrame(df_saldo)
        def cor_situacao(val):
            if val == "Devendo horas":
                return "color: #ff0000; font-weight: bold"
            elif val == "Horas extras":
                return "color: #00cc00; font-weight: bold"
            return ""
        st.dataframe(df_saldo_df.style.map(cor_situacao, subset=["Situação"]), hide_index=True, use_container_width=True)

    with tab2:
        if df_problemas:
            st.warning(f"{len(df_problemas)} registro(s) com batidas de ponto faltando")
            st.dataframe(pd.DataFrame(df_problemas), hide_index=True, use_container_width=True)
            for p in df_problemas:
                st.markdown(f"- **{p['Nome']}** em **{p['Data']}**: faltou **{p['Campos faltando']}**")
        else:
            st.success("Nenhum registro incompleto encontrado!")

    with tab3:
        st.subheader("Colaboradores com saldo negativo (devendo horas)")
        negativos = [s for s in df_saldo if s["Situação"] == "Devendo horas"]
        if negativos:
            st.dataframe(pd.DataFrame(negativos), hide_index=True, use_container_width=True)
        else:
            st.success("Nenhum colaborador devendo horas")

        st.subheader("Colaboradores com saldo positivo (horas extras)")
        positivos = [s for s in df_saldo if s["Situação"] == "Horas extras"]
        if positivos:
            st.dataframe(pd.DataFrame(positivos), hide_index=True, use_container_width=True)
        else:
            st.info("Nenhum colaborador com horas extras")

    with tab4:
        colaborador_sel = st.selectbox(
            "Selecionar colaborador",
            options=[(cid, info["nome"]) for cid, info in colaboradores.items()],
            format_func=lambda x: x[1],
        )
        if colaborador_sel:
            cid_sel = colaborador_sel[0]
            info = colaboradores[cid_sel]
            saldo_total = sum(d["saldo"] for d in info["dias"])
            st.subheader(f"{info['nome']} — {info['cargo']}")
            cols = st.columns(3)
            cols[0].metric("Carga horária diária", f"{info['carga_diaria']}h")
            cols[1].metric("Saldo total", fmt_saldo(saldo_total))
            cols[2].metric("Dias com registros", f"{len(info['dias'])}")
            st.dataframe(
                pd.DataFrame([{
                    "Data": d["data"],
                    "Entrada": fmt_hora(d["entrada"]),
                    "Saída Almoço": fmt_hora(d["saida_almoco"]),
                    "Retorno Almoço": fmt_hora(d["retorno_almoco"]),
                    "Saída": fmt_hora(d["saida"]),
                    "Horas Trab.": d["horas_trabalhadas"],
                    "Saldo": fmt_saldo(d["saldo"]),
                    "Problemas": ", ".join(d["problemas"]) if d["problemas"] else "OK",
                } for d in info["dias"]]),
                hide_index=True,
                use_container_width=True,
            )

if st.session_state.enviar_email:
    st.markdown("---")
    st.header("✉️ Envio de Alertas por E-mail")
    pendentes = {}
    for cid, info in colaboradores.items():
        saldo_total = sum(d["saldo"] for d in info["dias"])
        dias_problema = [(d["data"], d["problemas"]) for d in info["dias"] if d["problemas"]]
        if dias_problema or saldo_total < 0:
            pendentes[cid] = info

    if not pendentes:
        st.success("Nenhum colaborador com pendência encontrado!")
        st.session_state.enviar_email = False
    else:
        st.info(f"{len(pendentes)} colaborador(es) com pendência(s)")
        with st.spinner("Gerando e-mails com IA e enviando..."):
            for cid, info in pendentes.items():
                nome = info["nome"]
                email = info["email"]
                saldo_total = sum(d["saldo"] for d in info["dias"])
                dias_problema_lista = [(d["data"], d["problemas"]) for d in info["dias"] if d["problemas"]]
                if dias_problema_lista:
                    dias_str = "; ".join(f"{data} ({', '.join(probs)})" for data, probs in dias_problema_lista)
                    pendencia = f"Registros incompletos em: {dias_str}"
                else:
                    dias_str = ""
                    pendencia = f"Saldo negativo de horas: {fmt_saldo(saldo_total)}"
                st.markdown(f"**{nome}** ({email}) — {pendencia}")
                email_gerado = gerar_email_ia(nome, pendencia, dias_str, fmt_saldo(saldo_total), info["cargo"], tipo_ia, modelo_ia, api_key_ia, endpoint_ia)
                if email_gerado.startswith("Erro"):
                    st.error(f"Falha ao gerar e-mail para {nome}: {email_gerado}")
                else:
                    st.text_area(f"E-mail gerado para {nome}", email_gerado, height=200)
                    if usar_smtp and smtp_host and smtp_user and smtp_pass and smtp_dest:
                        assunto = "Alerta de pendência na folha de ponto"
                        corpo = email_gerado
                        linhas = email_gerado.split("\n")
                        for i, linha in enumerate(linhas):
                            if linha.lower().startswith("assunto"):
                                _, val = linha.split(":", 1)
                                assunto = val.strip()
                                corpo = "\n".join(linhas[i+1:])
                                break
                        sucesso, erro = enviar_email_smtp(assunto, corpo, email, smtp_host, smtp_port, smtp_user, smtp_pass, smtp_dest)
                        if sucesso:
                            st.success(f"E-mail enviado para {nome} ({email})")
                        else:
                            st.error(f"Falha ao enviar para {nome}: {erro}")
                    else:
                        st.info("E-mail exibido acima (SMTP não configurado). Configure SMTP na barra lateral para envio real.")
                st.divider()
        st.session_state.enviar_email = False

if st.session_state.analisar_ia:
    st.markdown("---")
    st.header(f"🤖 Análise com IA ({tipo_ia})")
    with st.spinner(f"Consultando {tipo_ia} ({modelo_ia})..."):
        resumo = gerar_resumo_texto(colaboradores)
        prompt = f"""Voce e um analista de RH. Analise os seguintes dados de folha de ponto e identifique:
1. Quais colaboradores tem dias com registros de ponto faltando (nao bateram o ponto em determinados horarios)
2. Quem esta com saldo de horas negativo (devendo horas) e quem esta com saldo positivo (horas extras)
3. Recomendacoes para cada caso

Dados:
{resumo}

Responda de forma clara e concisa em portugues."""
        resposta = consultar_ia(prompt, tipo_ia, modelo_ia, api_key_ia, endpoint_ia)
        st.markdown(resposta)
